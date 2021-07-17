

import configparser as cparser
import os
import shutil
import sys

import openpyxl
from config import setting
from openpyxl import load_workbook
from openpyxl.styles import Font  # 导入字体模块
from openpyxl.styles import PatternFill  # 导入填充模块

sys.path.append(os.path.dirname(os.path.dirname(__file__)))


# --------- 读取config.ini配置文件 ---------------
cf = cparser.ConfigParser()
cf.read(setting.TEST_CONFIG, encoding='UTF-8')
name = cf.get("tester", "name")


class ExcelHelper():
    '''
    操作读写excel
    '''

    def __init__(self, fileName, SheetName="Sheet1"):
        self.workbook = load_workbook(fileName)
        self.sheet = self.workbook[SheetName]
        self.fileName = fileName
        if not os.path.exists(setting.TARGET_FILE):
            # 文件不存在，则拷贝模板文件至指定报告目录下
            shutil.copyfile(setting.SOURCE_FILE, setting.TARGET_FILE)

        # 获取总行数、总列数
        self.max_row = self.sheet.max_row
        self.max_cow = self.sheet.max_column

    def write(self, row, value):
        """
        写入测试结果
        :param row_n:数据所在行数
        :param value: 测试结果值
        :return: 无
        """
        green = ['c6efce', '006100']  # 绿
        red = ['ffc7ce', '9c0006']  # 红

        if value == "PASS":
            fille = PatternFill('solid', fgColor=green[0])  # 设置填充颜色为
            font = Font(u'宋体', size=11, bold=True, italic=False,
                        strike=False, color=green[1])  # 设置字体样式

        if value == "FAIL":
            fille = PatternFill('solid', fgColor=red[0])
            font = Font(u'宋体', size=11, bold=True, italic=False,
                        strike=False, color=red[1])

        self.sheet.cell(row, 12, "").fill = fille
        self.sheet.cell(row, 12, value).font = font
        self.sheet.cell(row, 13, name)
        self.workbook.save(self.fileName)

    def read(self):
        """读取excel文件数据"""
        if self.max_row > 1:
            # 获取第一行的内容，列表格式
            keys = self.rows_values(1)
            listApiData = []
            # 获取每一行的内容，列表格式
            for col in range(2, self.max_row+1):
                values = self.rows_values(col)
                # keys，values组合转换为字典
                api_dict = dict(zip(keys, values))
                listApiData.append(api_dict)
            return listApiData
        else:
            print("表格是空数据!")
            return None

    def rows_values(self, row):
        '''
        获取某一行的内容
        '''
        row_list = []
        for i in self.sheet[row]:
            if i.value == None:
                row_list.append('')
            else:
                row_list.append(i.value)
        return row_list


if __name__ == "__main__":
    print(ExcelHelper("DemoAPITestCase.xlsx").read())
    #ExcelHelper("PersonApiTestCase.xlsx").write(2, "FAIL")
